# Recuperar datos de excel

def import_excel(ws, ini_cell, fin_cell):
    key_range = "%s:%s" %(ini_cell, fin_cell)
    
    cells = ws[key_range]

    datos = []

    for row in cells:
        p = {
            "Nombre": row[0].value,
            "Edad": row[1].value,
            "Sexo": row[2].value,
            "Categoria": row[3].value,
            "Salario": row[4].value
        }

        datos.append(p)

    return datos

from openpyxl import load_workbook

wb = load_workbook("data.xlsx", data_only=True)

ws = wb.active

datos = import_excel(ws, "A3", "E10")

print datos