# -*- coding: utf-8 -*-

def suma(valores):
    s = 0
    for x in valores:
        s = s + x
    return s

def suma2(valores):
    s = 0
    for x in valores:
        s = s + x ** 2
    return s

def promedio(valores):
    s = suma(valores)
    n = len(valores)
    return float(s) / n

def varianza(valores):
    p = promedio(valores)
    
    s = 0

    for x in valores:
        s = s + (x - p) ** 2

    n = len(valores)

    return float(s) / (n - 1)

def linspace(a, b, n):
    d = float(b - a) / (n - 1)
    
    l = []

    for i in range(n):
        l.append(a + i * d)

    return l

def bloque(ws, col, a, b):
    l = []

    for i in range(a, b + 1):
        c = ws["%s%d" %(col, i)]
        l.append(c.value if c.value != None else 0)

    return l

def inyectar(ws, col, row, values):
    n = len(values)
    for i in range(row, row + n):
        ws["%s%d" %(col, i)] = values[i - row]

def load_data(ws, ini_cell, fin_cell, labels):
    key_range = "%s:%s" %(ini_cell, fin_cell)
    
    cells = ws[key_range]

    datos = []

    for row in cells:
        p = {}

        for i in range(0, len(row)):
            p[labels[i]] = row[i].value

        datos.append(p)

    return datos

from openpyxl.utils import column_index_from_string

def get_labels(ws, ini_cell):
    cell = ws[ini_cell]

    labels = []

    while cell.value != None:
        labels.append(cell.value)

        cell = ws.cell(
            row = cell.row,
            column = column_index_from_string(cell.column) + 1
        )

    return labels

def get_max_column(ws, ini_cell):
    cell = ws[ini_cell]

    while cell.value != None:
        valid_cell = cell

        cell = ws.cell(
            row = cell.row,
            column = column_index_from_string(cell.column) + 1
        )

    return valid_cell.column + str(valid_cell.row)

def get_max_row(ws, ini_cell):
    cell = ws[ini_cell]

    while cell.value != None:
        valid_cell = cell

        cell = ws.cell(
            row = cell.row + 1,
            column = column_index_from_string(cell.column)
        )

    return valid_cell.column + str(valid_cell.row)