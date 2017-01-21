from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws["A1"] = "Hola mundo"

ws["B5"] = 123

ws["AA1"] = "gg"

wb.save("mi_libro.xlsx")

