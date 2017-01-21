from openpyxl import Workbook

X = [1, 3, 5, 7, 9, 11]
Y = [9, 22, 67, 54, 38, 99]

n = len(X)

wb = Workbook()

ws = wb.active

for i in range(0, n):
    ws["A%d" %(i + 1)] = X[i]
    ws["B%d" %(i + 1)] = Y[i]

wb.save("mi_libro.xlsx")

