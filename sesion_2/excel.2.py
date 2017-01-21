from openpyxl import Workbook
from geg import linspace
import math

n = 10

X = linspace(-math.pi, math.pi, n)

wb = Workbook()

ws = wb.active

for i in range(0, n):
    x = X[i]

    ws["A%d" %(i + 1)] = x
    ws["B%d" %(i + 1)] = math.sin(x)
    ws["C%d" %(i + 1)] = math.cos(x)
    ws["D%d" %(i + 1)] = math.tan(x)

wb.save("sin_cos_tan.xlsx")

