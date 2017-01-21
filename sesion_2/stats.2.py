from openpyxl import load_workbook
from geg import *

wb = load_workbook("data.xlsx", data_only=True)

ws = wb.active

rango = ws["A1:E10"]

def transformacion(cell):
    if cell.column == "C":
        if cell.value == "H" or cell.value == "Hombre":
            return 1
        return 2

    return cell.value

for row in rango:
    print map(transformacion, row)