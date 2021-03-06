def load_mat(cells):
    return [([cell.value for cell in xrow]) for xrow in cells]

def data_map(data, fn):
    aux = []

    for x in data:
        y = fn(x) # Pasa el filtro o transformacion?
        if type(y) == bool: # fn es filtro
            if y:
                aux.append(x)
        elif y != None: # fn es transformacion
            aux.append(y)
    
    return aux

def data_count(data, fn):
    return len(data_map(data, fn))

def data_min(data, labels):
    return data_map(data, lambda l: labels[l.index(min(l))])

def data_max(data, labels):
    return data_map(data, lambda l: labels[l.index(max(l))])

def load_data(cells):
    mat = load_mat(cells)

    labels = mat[0]

    data = []

    for i in range(1, len(mat)):
        row = mat[i]

        dic = {}
        for k, v in zip(labels, row):
            dic[k] = v

        data.append(dic)

    return labels, data

def plot_pie(ax, data, label):
    cats = {}

    for dic in data:
        k = dic[label]
        if cats.has_key(k):
            cats[k] += 1
        else:
            cats[k] = 1

    labels, sizes = zip(*[(k, cats[k]) for k in cats])

    ax.pie(sizes, labels=labels, autopct="%.0f%%")
    ax.axis('equal')

def data_append(data, label, fn):
    for dic in data:
        dic[label] = fn(dic)
    return data

def data_cat(data, label, funs):
    for dic in data:
        value = "None"
        for cat in funs:
            fn = funs[cat]
            if fn(dic):
                value = cat
        dic[label] = value
    return data 

from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

def save_data(data, labels, filename, sheet_name, ini_cell):
    try:
        wb = load_workbook(filename)
    except:
        wb = Workbook()

    if wb.sheetnames.count(sheet_name) > 0:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    for j in range(len(labels)):
        cell = ws.cell(
            row = ws[ini_cell].row,
            column = column_index_from_string(ws[ini_cell].column) + j
        )
        cell.value = labels[j]
    for i in range(len(data)):
        dic = data[i]
        for j in range(len(labels)):
            cell = ws.cell(
                row = ws[ini_cell].row + 1 + i,
                column = column_index_from_string(ws[ini_cell].column) + j
            )
            cell.value = dic[labels[j]]
    wb.save(filename)

def data_column(data, column):
    return [dic[column] for dic in data]

def data_row(data, index, labels):
    return zip(*[(k, data[index][k]) for k in labels])